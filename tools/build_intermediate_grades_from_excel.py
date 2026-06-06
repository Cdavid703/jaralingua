#!/usr/bin/env python3
import argparse
import datetime as dt
import json
import re
from pathlib import Path

import openpyxl


EVALUATION_IDS = {
    "MIDTERM WRITING TASK (20%)": ("midtermWritingTask", "Writing"),
    "MIDTERM ORAL TASK (20%)": ("midtermOralTask", "Speaking"),
}


def clean_id(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value))
    return re.sub(r"\D", "", str(value))


def clean_name(value):
    text = " ".join(str(value or "").strip().split())
    return re.sub(r"^\d+\.\s*", "", text).strip()


def clean_email(value):
    return str(value or "").strip().lower()


def slugify(value):
    text = re.sub(r"[^A-Za-z0-9]+", " ", str(value or "")).strip().title().replace(" ", "")
    return (text[:1].lower() + text[1:]) if text else "assessment"


def weight_from_title(title):
    match = re.search(r"\((\d+(?:\.\d+)?)%\)", title)
    if not match:
        return 0
    weight = float(match.group(1))
    return int(weight) if weight.is_integer() else weight


def grade_value(value):
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        value = float(f"{value.day}.{value.month}")
    elif isinstance(value, dt.date):
        value = float(f"{value.day}.{value.month}")
    else:
        try:
            value = float(str(value).replace(",", "."))
        except ValueError:
            return None
    if value < 0 or value > 5:
        return None
    return round(value, 2)


def build_payload(source_path):
    workbook = openpyxl.load_workbook(source_path, data_only=False)
    roster_sheet = workbook["Hoja 1"]
    notes_sheet = workbook["Notas"]

    roster = {}
    for row in roster_sheet.iter_rows(min_row=2, values_only=True):
        student_id = clean_id(row[1] if len(row) > 1 else "")
        if not student_id:
            continue
        roster[student_id] = {
            "id": student_id,
            "fullName": clean_name(row[0] if len(row) > 0 else ""),
            "email": clean_email(row[2] if len(row) > 2 else ""),
        }

    evaluations = []
    for column_index in range(3, notes_sheet.max_column + 1):
        title = notes_sheet.cell(1, column_index).value
        if not title:
            continue
        title = str(title).strip()
        evaluation_id, evaluation_type = EVALUATION_IDS.get(title, (slugify(title), "Assessment"))
        evaluations.append({
            "column": column_index,
            "id": evaluation_id,
            "title": title,
            "weight": weight_from_title(title),
            "type": evaluation_type,
            "description": title,
        })

    students = []
    for row_index in range(2, notes_sheet.max_row + 1):
        student_id = clean_id(notes_sheet.cell(row_index, 1).value)
        if not student_id:
            continue
        roster_record = roster.get(student_id, {})
        full_name = roster_record.get("fullName") or clean_name(notes_sheet.cell(row_index, 2).value)
        grades = {}
        for evaluation in evaluations:
            grade = grade_value(notes_sheet.cell(row_index, evaluation["column"]).value)
            if grade is not None:
                grades[evaluation["id"]] = grade
        students.append({
            "id": student_id,
            "fullName": full_name,
            "level": "Intermediate English Course 1",
            "email": roster_record.get("email", ""),
            "emailAliases": [],
            "contact": "",
            "bookDate": None,
            "grades": grades,
        })

    return {
        "adminEmails": ["cdavid.jaramillo@gmail.com"],
        "teacherEmails": [],
        "allowStudentIdClaim": False,
        "students": students,
        "evaluations": [{key: value for key, value in item.items() if key != "column"} for item in evaluations],
        "bonusEvent": None,
    }


def main():
    parser = argparse.ArgumentParser(description="Build private Intermediate English grades JSON from workbook.")
    parser.add_argument("source", help="Path to source .xlsx workbook.")
    parser.add_argument("--out", default="data/intermediate-english-grades.local.json", help="Output JSON path.")
    args = parser.parse_args()

    payload = build_payload(args.source)
    output_path = Path(args.out)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    missing_emails = [student["fullName"] for student in payload["students"] if not student.get("email")]
    print(json.dumps({
        "students": len(payload["students"]),
        "evaluations": len(payload["evaluations"]),
        "missingEmails": missing_emails,
        "output": str(output_path),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
